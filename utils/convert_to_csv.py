import ast
import csv
from openpyxl import Workbook
from pathlib import Path
import os

import logging
logger = logging.getLogger("CSV")
logger.setLevel(level=logging.INFO)

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers
from datetime import datetime
from openpyxl.styles import Alignment

def convert_to_csv(input_file, output_file):
    rows = []

    with open(input_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                data = ast.literal_eval(line)

                rows.append(data)

    # Write to CSV
    fieldnames = ['TIN', 'Total', 'Date Issued', 'Company & Address', 'Link', 'Image Path']

    with open(output_file, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()

        for row in rows:
            # Filter only keys present in fieldnames
            filtered_row = {k: v for k, v in row.items() if k in fieldnames}
            writer.writerow(filtered_row)

    convert_to_xlsx(rows, "output/final_report.xlsx")

    logger.info(f"Converted to csv successfully! {output_file}")


def fix_date(date_str):
    """
    Fix invalid year formats like 2609-01-06 -> 2026-01-06
    """
    try:
        if len(date_str) >= 10:
            year = date_str[:4]

            # If year is weird like 2609 #Happens when 01/01/2619:00:0000
            if int(year) > 2100:
                # assume last 2 digits are correct year
                corrected_year = "20" + year[:2]
                date_str = corrected_year + date_str[4:]

        return datetime.strptime(date_str[:10], "%Y-%m-%d")

    except:
        return None


def clean_total(value):
    """
    Convert 1,000.00 or 1168 to float
    """
    try:
        return float(str(value).replace(",", ""))
    except:
        return None


def convert_to_xlsx(rows, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts"

    headers = [
        'TIN', 'Total', 'Date Issued',
        'Company & Address', 'Link',
        'Image Path'
    ]

    ws.append(headers)

    for row in rows:
        cleaned_row = []

        for h in headers:
            value = row.get(h, "")

            if h == "Total":
                value = clean_total(value)

            elif h == "Date Issued":
                value = fix_date(value)

            cleaned_row.append(value)

        ws.append(cleaned_row)

        current_row = ws.max_row

        # Format Total column as currency
        total_col = headers.index("Total") + 1
        ws.cell(row=current_row, column=total_col).number_format = '#,##0.00'

        # Format Date column
        date_col = headers.index("Date Issued") + 1
        ws.cell(row=current_row, column=date_col).number_format = 'yyyy-mm-dd'

        # Hyperlink for Image Path
        img_col = headers.index("Image Path") + 1
        image_path_str = row.get("Image Path", "")

        if image_path_str:
            filename = Path(image_path_str).name
            correct_path = f"../receipts_v2/{filename}"

            cell = ws.cell(row=current_row, column=img_col)
            cell.value = "Open Receipt"
            cell.hyperlink = correct_path
            cell.style = "Hyperlink"

    # ---- CREATE REAL EXCEL TABLE ----
    table_end_row = ws.max_row
    table_ref = f"A1:F{table_end_row}"

    table = Table(displayName="ReceiptsTable", ref=table_ref)

    style = TableStyleInfo(
        name="TableStyleMedium1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    # Optional: Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(wrap_text=True)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file)